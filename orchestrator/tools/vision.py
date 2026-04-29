"""
orchestrator/tools/vision.py — Visual perception tools for agents.

Gives agents the ability to "see" — converting files and URLs into images
that can be passed to vision-capable LLMs as multimodal message content.

Tools:
  read_image        — Load a workspace image file for the agent to see
  read_pdf          — Render PDF pages as images (via PyMuPDF)
  read_spreadsheet  — Render Excel sheets as a table image (via openpyxl + pillow)
  screenshot_url    — Screenshot a webpage (via playwright, optional)
  analyze_image     — Ask a vision model to describe/analyze an image

All tools return base64-encoded PNG images that get injected into the
next LLM message as image_url blocks.  The agent literally sees the content.

Provider vision support:
  ✓ OpenAI GPT-4V, GPT-4o — full vision
  ✓ Anthropic Claude 3/4 — full vision
  ✓ Gemini 1.5 Flash/Pro — full vision
  ✓ Ollama llava, bakllava, moondream — local vision
  ✓ Groq llava-v1.5-7b — vision (limited)
  ✗ Most text-only models — images stripped, text description injected instead
"""

from __future__ import annotations

import base64
import io
from pathlib import Path
from typing import Any

import structlog

from orchestrator.tools.file_store import project_dir, ResourceType

log = structlog.get_logger(__name__)

_MAX_IMAGE_BYTES = 5 * 1024 * 1024   # 5MB per image
_MAX_PDF_PAGES   = 8                  # don't flood context with huge PDFs
_MAX_SHEET_ROWS  = 50                 # rows to render in spreadsheet preview


# ---------------------------------------------------------------------------
# read_image
# ---------------------------------------------------------------------------

async def read_image(
    path: str,
    project_id: str,
    max_width: int = 1024,
) -> dict[str, Any]:
    """
    Load an image from the project workspace and return it as a base64 block
    ready to inject into a multimodal message.
    """
    safe = Path(path).as_posix().lstrip("/")
    if ".." in safe:
        return _err("read_image", "Invalid path: directory traversal not allowed")

    full = project_dir(project_id) / safe
    if not full.exists():
        return _err("read_image", f"File not found: {safe}")

    try:
        from PIL import Image as PILImage
        img = PILImage.open(full)
        if img.width > max_width:
            ratio = max_width / img.width
            img = img.resize((max_width, int(img.height * ratio)), PILImage.LANCZOS)
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGB")

        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        b64 = base64.b64encode(buf.getvalue()).decode()
    except Exception as exc:
        return _err("read_image", f"Failed to read image: {exc}")

    return {
        "tool": "read_image",
        "success": True,
        "path": safe,
        "width": img.width,
        "height": img.height,
        "visual_blocks": [_img_block(b64, f"Image: {Path(safe).name}")],
        "files_created": [],
        "description": f"Image {Path(safe).name} ({img.width}×{img.height}px) loaded for visual inspection.",
    }


# ---------------------------------------------------------------------------
# read_pdf
# ---------------------------------------------------------------------------

async def read_pdf(
    path: str,
    project_id: str,
    pages: str = "1-3",
    max_width: int = 900,
) -> dict[str, Any]:
    """
    Render PDF pages as images for the agent to read visually.
    pages: "1", "1-3", "all" (capped at _MAX_PDF_PAGES)
    """
    safe = Path(path).as_posix().lstrip("/")
    if ".." in safe:
        return _err("read_pdf", "Invalid path")

    full = project_dir(project_id) / safe
    if not full.exists():
        # Also check raw documents dir
        full = project_dir(project_id) / "documents" / safe
        if not full.exists():
            return _err("read_pdf", f"File not found: {safe}")

    try:
        import fitz   # PyMuPDF
    except ImportError:
        return _err("read_pdf", "PyMuPDF not installed. Run: pip install pymupdf")

    try:
        doc = fitz.open(str(full))
        total_pages = len(doc)

        # Parse page range
        page_nums = _parse_page_range(pages, total_pages)

        visual_blocks = []
        for pnum in page_nums:
            page = doc[pnum]
            # Render at 1.5x zoom for readability
            mat  = fitz.Matrix(1.5, 1.5)
            pix  = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            img_bytes = pix.tobytes("png")

            # Resize if too wide
            if pix.width > max_width:
                from PIL import Image as PILImage
                img = PILImage.open(io.BytesIO(img_bytes))
                ratio = max_width / img.width
                img = img.resize((max_width, int(img.height * ratio)), PILImage.LANCZOS)
                buf = io.BytesIO()
                img.save(buf, format="PNG")
                img_bytes = buf.getvalue()

            b64 = base64.b64encode(img_bytes).decode()
            visual_blocks.append(_img_block(b64, f"PDF page {pnum + 1} of {total_pages}"))

        # Also extract text for models without vision
        text_preview = ""
        for pnum in page_nums[:3]:
            text_preview += doc[pnum].get_text()[:1000]

        doc.close()

    except Exception as exc:
        return _err("read_pdf", f"Failed to render PDF: {exc}")

    return {
        "tool": "read_pdf",
        "success": True,
        "path": safe,
        "total_pages": total_pages,
        "pages_shown": [p + 1 for p in page_nums],
        "visual_blocks": visual_blocks,
        "text_preview": text_preview[:2000],
        "files_created": [],
        "description": (
            f"PDF '{Path(safe).name}' — {total_pages} pages total. "
            f"Showing pages {[p+1 for p in page_nums]}."
        ),
    }


# ---------------------------------------------------------------------------
# read_spreadsheet
# ---------------------------------------------------------------------------

async def read_spreadsheet(
    path: str,
    project_id: str,
    sheet: str | None = None,
    max_rows: int = _MAX_SHEET_ROWS,
) -> dict[str, Any]:
    """
    Render an Excel spreadsheet as a table image for the agent to inspect.
    Also returns the raw data for text-only models.
    """
    safe = Path(path).as_posix().lstrip("/")
    if ".." in safe:
        return _err("read_spreadsheet", "Invalid path")

    full = project_dir(project_id) / safe
    if not full.exists():
        full = project_dir(project_id) / "data" / safe
        if not full.exists():
            return _err("read_spreadsheet", f"File not found: {safe}")

    try:
        import openpyxl
    except ImportError:
        return _err("read_spreadsheet", "openpyxl not installed. Run: pip install openpyxl")

    try:
        wb = openpyxl.load_workbook(full, read_only=True, data_only=True)
        sheet_name = sheet or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]

        rows_data = []
        for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
            if any(c is not None for c in row):
                rows_data.append([str(c or "") for c in row])
        wb.close()

        if not rows_data:
            return _err("read_spreadsheet", "Spreadsheet appears empty")

        # Render as image table
        img_bytes = _render_table_image(rows_data, title=f"{Path(safe).name} — {sheet_name}")
        b64 = base64.b64encode(img_bytes).decode()

        # Text representation for fallback
        text_repr = "\n".join("\t".join(row) for row in rows_data[:20])

        return {
            "tool": "read_spreadsheet",
            "success": True,
            "path": safe,
            "sheet": sheet_name,
            "rows": len(rows_data),
            "cols": max(len(r) for r in rows_data) if rows_data else 0,
            "visual_blocks": [_img_block(b64, f"Spreadsheet: {Path(safe).name}")],
            "text_data": text_repr,
            "files_created": [],
            "description": f"Spreadsheet '{sheet_name}' with {len(rows_data)} rows rendered for inspection.",
        }

    except Exception as exc:
        return _err("read_spreadsheet", f"Failed to render spreadsheet: {exc}")


def _render_table_image(rows: list[list[str]], title: str = "") -> bytes:
    """Render tabular data as a clean PNG image."""
    from PIL import Image as PILImage, ImageDraw, ImageFont

    # Layout
    CELL_PAD_X = 12
    CELL_PAD_Y = 8
    FONT_SIZE  = 13
    HEADER_BG  = (30, 58, 95)
    ALT_ROW_BG = (240, 244, 248)
    BORDER_COL = (200, 210, 220)
    TEXT_COL   = (30, 40, 60)
    HDR_TEXT   = (255, 255, 255)
    BG_COL     = (255, 255, 255)

    # Attempt to load a monospace font; fall back to default
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf", FONT_SIZE)
        hdr_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSansMono-Bold.ttf", FONT_SIZE)
    except Exception:
        font = ImageFont.load_default()
        hdr_font = font

    # Measure column widths using a dummy image
    dummy = PILImage.new("RGB", (1, 1))
    dummy_draw = ImageDraw.Draw(dummy)

    col_widths = []
    num_cols = max(len(r) for r in rows)
    for col in range(num_cols):
        max_w = 0
        for row in rows:
            if col < len(row):
                w = dummy_draw.textlength(str(row[col])[:40], font=font)
                max_w = max(max_w, w)
        col_widths.append(int(max_w) + CELL_PAD_X * 2)

    row_height = FONT_SIZE + CELL_PAD_Y * 2
    img_width  = max(sum(col_widths) + 2, 400)
    img_height = row_height * len(rows) + (30 if title else 0) + 4

    img  = PILImage.new("RGB", (img_width, img_height), BG_COL)
    draw = ImageDraw.Draw(img)

    y_off = 0
    if title:
        draw.rectangle([0, 0, img_width, 28], fill=(15, 30, 60))
        draw.text((8, 6), title[:80], fill=(200, 220, 255), font=hdr_font)
        y_off = 30

    for row_i, row in enumerate(rows):
        x = 0
        bg = HEADER_BG if row_i == 0 else (ALT_ROW_BG if row_i % 2 == 0 else BG_COL)
        draw.rectangle([0, y_off, img_width, y_off + row_height], fill=bg)
        for col_i, cell in enumerate(row):
            if col_i >= len(col_widths):
                break
            col_w = col_widths[col_i]
            text  = str(cell)[:40]
            fg    = HDR_TEXT if row_i == 0 else TEXT_COL
            f     = hdr_font if row_i == 0 else font
            draw.text((x + CELL_PAD_X, y_off + CELL_PAD_Y), text, fill=fg, font=f)
            draw.line([(x + col_w, y_off), (x + col_w, y_off + row_height)], fill=BORDER_COL)
            x += col_w
        draw.line([(0, y_off + row_height), (img_width, y_off + row_height)], fill=BORDER_COL)
        y_off += row_height

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# screenshot_url
# ---------------------------------------------------------------------------

async def screenshot_url(
    url: str,
    project_id: str,
    wait_ms: int = 1500,
    max_width: int = 1200,
) -> dict[str, Any]:
    """
    Take a screenshot of a URL using Playwright.
    Requires: pip install playwright && playwright install chromium
    """
    if not url.startswith(("http://", "https://")):
        return _err("screenshot_url", f"URL must start with http:// or https://: {url!r}")

    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return _err(
            "screenshot_url",
            "Playwright not installed. Run: pip install playwright && playwright install chromium"
        )

    try:
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True)
            page    = await browser.new_page(viewport={"width": max_width, "height": 800})
            await page.goto(url, wait_until="networkidle", timeout=15000)
            if wait_ms > 0:
                await page.wait_for_timeout(wait_ms)
            img_bytes = await page.screenshot(full_page=False, type="png")
            await browser.close()
    except Exception as exc:
        return _err("screenshot_url", f"Screenshot failed: {exc}")

    b64 = base64.b64encode(img_bytes).decode()

    return {
        "tool": "screenshot_url",
        "success": True,
        "url": url,
        "visual_blocks": [_img_block(b64, f"Screenshot of {url}")],
        "files_created": [],
        "description": f"Screenshot of {url} captured for visual inspection.",
    }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _img_block(b64: str, alt: str = "") -> dict:
    return {
        "type": "image_url",
        "image_url": {"url": f"data:image/png;base64,{b64}"},
        "_alt": alt,   # informational only, not sent to API
    }


def _err(tool: str, msg: str) -> dict:
    return {"tool": tool, "success": False, "error": msg,
            "visual_blocks": [], "files_created": []}


def _parse_page_range(pages: str, total: int) -> list[int]:
    """Parse page spec like '1', '1-3', 'all' → 0-indexed list."""
    pages = pages.strip().lower()
    if pages == "all":
        return list(range(min(total, _MAX_PDF_PAGES)))
    if "-" in pages:
        parts = pages.split("-")
        start = max(0, int(parts[0]) - 1)
        end   = min(total, int(parts[1]))
        return list(range(start, min(end, start + _MAX_PDF_PAGES)))
    try:
        n = int(pages) - 1
        return [max(0, min(n, total - 1))]
    except ValueError:
        return [0]
