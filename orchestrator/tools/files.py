"""
orchestrator/tools/files.py — File creation and reading tools.

create_spreadsheet: openpyxl → .xlsx
create_document:    plain text → .md or .txt
read_file:          reads any file from project workspace
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import structlog

from orchestrator.tools.file_store import (
    project_dir, save_bytes, save_text, unique_filename,
    ResourceType, _EXT_TO_TYPE,
)

log = structlog.get_logger(__name__)


# ---------------------------------------------------------------------------
# create_spreadsheet
# ---------------------------------------------------------------------------

async def create_spreadsheet(
    sheets: list[dict],
    filename: str,
    project_id: str,
) -> dict[str, Any]:
    """
    Create an Excel spreadsheet with one or more sheets.

    sheets format:
        [
          {
            "name": "Sheet1",
            "headers": ["Col A", "Col B", "Col C"],
            "rows": [
              ["val1", "val2", "val3"],
              ...
            ]
          }
        ]
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {
            "tool": "create_spreadsheet",
            "success": False,
            "error": "openpyxl not installed. Run: pip install openpyxl",
            "files_created": [],
        }

    if not sheets:
        return {
            "tool": "create_spreadsheet",
            "success": False,
            "error": "No sheets provided",
            "files_created": [],
        }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # Styles
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1e3a5f")
    alt_row_fill   = PatternFill("solid", fgColor="f0f4f8")
    center_align   = Alignment(horizontal="center", vertical="center")
    thin_border    = Border(
        left=Side(style="thin", color="d0d8e4"),
        right=Side(style="thin", color="d0d8e4"),
        top=Side(style="thin", color="d0d8e4"),
        bottom=Side(style="thin", color="d0d8e4"),
    )

    for sheet_def in sheets:
        name    = str(sheet_def.get("name", "Sheet"))[:31]  # Excel limit
        headers = sheet_def.get("headers", [])
        rows    = sheet_def.get("rows", [])

        ws = wb.create_sheet(title=name)

        # Write headers
        if headers:
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(header))
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = center_align
                cell.border    = thin_border

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_row_fill

        # Auto-width columns
        for col_idx in range(1, max(len(headers), 1) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        max_length = max(max_length, len(str(cell.value or "")))
                    except Exception:
                        pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

        # Freeze header row
        if headers:
            ws.freeze_panes = "A2"

    # Save to bytes
    import io
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    fname = unique_filename(f"{filename}.xlsx", project_id, ResourceType.DATA)
    dest  = save_bytes(project_id, fname, xlsx_bytes, ResourceType.DATA)
    rel   = str(dest.relative_to(project_dir(project_id)))

    total_rows = sum(len(s.get("rows", [])) for s in sheets)

    return {
        "tool": "create_spreadsheet",
        "success": True,
        "filename": fname,
        "path": rel,
        "sheets": len(sheets),
        "total_rows": total_rows,
        "files_created": [rel],
    }


# ---------------------------------------------------------------------------
# create_document
# ---------------------------------------------------------------------------

async def create_document(
    content: str,
    filename: str,
    project_id: str,
    format: str = "md",
) -> dict[str, Any]:
    """Create a text document (.md or .txt) and save to workspace."""
    ext = "md" if format.lower() == "md" else "txt"
    fname = unique_filename(f"{filename}.{ext}", project_id, ResourceType.DOCUMENT)
    dest  = save_text(project_id, fname, content, ResourceType.DOCUMENT)
    rel   = str(dest.relative_to(project_dir(project_id)))

    return {
        "tool": "create_document",
        "success": True,
        "filename": fname,
        "path": rel,
        "char_count": len(content),
        "line_count": content.count("\n") + 1,
        "files_created": [rel],
    }


# ---------------------------------------------------------------------------
# read_file
# ---------------------------------------------------------------------------

async def read_file(
    path: str,
    project_id: str,
    max_chars: int = 10_000,
) -> dict[str, Any]:
    """
    Read a file from the project workspace.
    Handles text files directly; returns metadata + preview for binary files.
    """
    # Sanitise path — prevent directory traversal
    safe_path = Path(path).as_posix().lstrip("/")
    if ".." in safe_path:
        return {
            "tool": "read_file",
            "success": False,
            "error": "Invalid path: directory traversal not allowed",
            "files_created": [],
        }

    full_path = project_dir(project_id) / safe_path
    if not full_path.exists():
        return {
            "tool": "read_file",
            "success": False,
            "error": f"File not found: {safe_path}",
            "files_created": [],
        }

    if not full_path.is_file():
        return {
            "tool": "read_file",
            "success": False,
            "error": f"Path is not a file: {safe_path}",
            "files_created": [],
        }

    ext = full_path.suffix.lower()
    size_bytes = full_path.stat().st_size

    # Binary file types — return metadata only
    binary_exts = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".mp3", ".wav",
                   ".ogg", ".mp4", ".pdf", ".xlsx", ".xls", ".docx", ".zip"}
    if ext in binary_exts:
        # For Excel files, try to extract text summary
        if ext == ".xlsx":
            return await _read_xlsx(full_path, project_id, safe_path, size_bytes)

        return {
            "tool": "read_file",
            "success": True,
            "path": safe_path,
            "type": "binary",
            "extension": ext,
            "size_bytes": size_bytes,
            "content": f"[Binary file: {ext}, {size_bytes:,} bytes. Cannot display as text.]",
            "files_created": [],
        }

    # Text file
    try:
        raw = full_path.read_text(encoding="utf-8", errors="replace")
        content = raw[:max_chars]
        truncated = len(raw) > max_chars

        return {
            "tool": "read_file",
            "success": True,
            "path": safe_path,
            "type": "text",
            "extension": ext,
            "size_bytes": size_bytes,
            "char_count": len(raw),
            "content": content,
            "truncated": truncated,
            "files_created": [],
        }
    except Exception as exc:
        return {
            "tool": "read_file",
            "success": False,
            "error": f"Failed to read file: {exc}",
            "files_created": [],
        }


async def _read_xlsx(
    path: Path, project_id: str, rel_path: str, size_bytes: int
) -> dict:
    """Extract a text summary from an Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        summary_lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            summary_lines.append(f"## Sheet: {sheet_name}")
            row_count = 0
            for row in ws.iter_rows(max_row=20, values_only=True):
                if any(c is not None for c in row):
                    summary_lines.append("\t".join(str(c or "") for c in row))
                    row_count += 1
            if row_count == 20:
                summary_lines.append("... (truncated at 20 rows per sheet)")
            summary_lines.append("")
        wb.close()
        content = "\n".join(summary_lines)

        return {
            "tool": "read_file",
            "success": True,
            "path": rel_path,
            "type": "spreadsheet",
            "sheets": wb.sheetnames,
            "size_bytes": size_bytes,
            "content": content,
            "files_created": [],
        }
    except Exception as exc:
        return {
            "tool": "read_file",
            "success": True,
            "path": rel_path,
            "type": "binary",
            "extension": ".xlsx",
            "size_bytes": size_bytes,
            "content": f"[Excel file: {size_bytes:,} bytes. Could not parse: {exc}]",
            "files_created": [],
        }
